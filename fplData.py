import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

print(f"Initialising...")
# === Set your mini-league ID ===
league_id = '542663'

# === Get league standings ===
league_url = f"https://fantasy.premierleague.com/api/leagues-classic/{league_id}/standings/"
response = requests.get(league_url)
league_data = response.json()

entry_map = {}
for entry in league_data['standings']['results']:
    entry_map[entry['entry']] = {
        'manager_name': entry['player_name'],
        'team_name': entry['entry_name']
    }

entries = list(entry_map.keys())

# === Get player static data ===
bootstrap = requests.get("https://fantasy.premierleague.com/api/bootstrap-static/").json()
elements = bootstrap['elements']
teams = bootstrap['teams']
positions = bootstrap['element_types']

id_to_name = {e['id']: e['web_name'] for e in elements}
id_to_team = {e['id']: teams[e['team'] - 1]['name'] for e in elements}
id_to_position = {e['id']: positions[e['element_type'] - 1]['singular_name'] for e in elements}

# === Step 1: Collect chip usage ===
chip_data = []
free_hit_by_entry = {}
triple_captain_week_by_entry = {}

print(f"Fetching chip data...")
for entry_id in entries:
    history_url = f"https://fantasy.premierleague.com/api/entry/{entry_id}/history/"
    resp = requests.get(history_url)

    chip_dict = {
        'Manager Name': entry_map[entry_id]['manager_name'],
        'Team Name': entry_map[entry_id]['team_name'],
        'Wildcard 1': '-',
        'Wildcard 2': '-',
        'Free Hit': '-',
        'Bench Boost': '-',
        'Triple Captain': '-',
        'Triple Captain Player': '-',
        'TC Points': '-',
    }

    free_hit_week = None
    triple_captain_week = None

    if resp.status_code == 200:
        chips = resp.json().get("chips", [])
        for chip in chips:
            name = chip['name']
            gw = chip['event']
            if name == "wildcard":
                if gw <= 20:
                    chip_dict["Wildcard 1"] = gw
                else:
                    chip_dict["Wildcard 2"] = gw
            elif name == "freehit":
                chip_dict["Free Hit"] = gw
                free_hit_week = gw
            elif name == "3xc":
                chip_dict["Triple Captain"] = gw
                triple_captain_week = gw
            elif name == "bboost":
                chip_dict["Bench Boost"] = gw

    free_hit_by_entry[entry_id] = free_hit_week
    triple_captain_week_by_entry[entry_id] = triple_captain_week

    if triple_captain_week:
        picks_url = f"https://fantasy.premierleague.com/api/entry/{entry_id}/event/{triple_captain_week}/picks/"
        live_url = f"https://fantasy.premierleague.com/api/event/{triple_captain_week}/live/"

        picks_resp = requests.get(picks_url)
        live_resp = requests.get(live_url)

        if picks_resp.status_code == 200 and live_resp.status_code == 200:
            captain_id = next((p['element'] for p in picks_resp.json()['picks'] if p['is_captain']), None)
            if captain_id:
                chip_dict["Triple Captain Player"] = id_to_name.get(captain_id, '-')

                # Get points from live data
                player_stats = {p['id']: p['stats']['total_points'] for p in live_resp.json()['elements']}
                chip_dict["TC Points"] = player_stats.get(captain_id, '-')

    chip_data.append(chip_dict)
print(f"✅")

df_chips = pd.DataFrame(chip_data)

# === Step 2: Optimized Captaincy Data Collection ===
captaincy_data = []

for gw in range(1, 39):
    print(f"Fetching GW{gw} captain data...")

    # Get player points once for the whole GW
    live_url = f"https://fantasy.premierleague.com/api/event/{gw}/live/"
    live_resp = requests.get(live_url)
    if live_resp.status_code != 200:
        print(f"⚠️ Skipping GW{gw} - live data unavailable.")
        continue

    player_points_map = {p['id']: p['stats']['total_points'] for p in live_resp.json()['elements']}

    for entry_id in entries:
        picks_url = f"https://fantasy.premierleague.com/api/entry/{entry_id}/event/{gw}/picks/"
        picks_resp = requests.get(picks_url)

        if picks_resp.status_code != 200:
            continue

        picks = picks_resp.json().get('picks', [])
        captain_id = next((p['element'] for p in picks if p['is_captain']), None)

        if captain_id:
            manager_name = entry_map[entry_id]['manager_name']
            team_name = entry_map[entry_id]['team_name']
            player_name = id_to_name.get(captain_id, '-')
            player_points = player_points_map.get(captain_id, '-')
            triple_captain_used = "Yes" if gw == triple_captain_week_by_entry.get(entry_id) else "No"

            # Add captaincy record
            captaincy_data.append({
                'Manager Name': manager_name,
                'Team Name': team_name,
                'Gameweek': gw,
                'Captain': player_name,
                'Captain Points': player_points,
                'Triple Captain Used': triple_captain_used
            })

print(f"✅")
df_captaincy = pd.DataFrame(captaincy_data)

# === Step 3: Collect transfer data ===
all_transfers = []

print(f"Fetching transfer data...")
for entry_id in entries:
    transfers_url = f"https://fantasy.premierleague.com/api/entry/{entry_id}/transfers/"
    resp = requests.get(transfers_url)

    if resp.status_code == 200:
        transfers = resp.json()
        for t in transfers:
            if t['element_in'] not in id_to_name or t['element_out'] not in id_to_name:
                continue

            free_hit_used = "Yes" if t['event'] == free_hit_by_entry.get(entry_id) else "No"

            all_transfers.append({
                'Manager Name': entry_map[entry_id]['manager_name'],
                'Team Name': entry_map[entry_id]['team_name'],
                'Gameweek': t['event'],
                'Player Out': id_to_name[t['element_out']],
                'Out - Team': id_to_team[t['element_out']],
                'Out - Position': id_to_position[t['element_out']],
                'Player In': id_to_name[t['element_in']],
                'In - Team': id_to_team[t['element_in']],
                'In - Position': id_to_position[t['element_in']],
                'Free Hit Active': free_hit_used
            })
print(f"✅")

df_transfers = pd.DataFrame(all_transfers)

if not df_transfers.empty:
    df_transfers.sort_values(by=['Gameweek', 'Manager Name'], inplace=True)

# === Step 4: Write to Excel ===
print(f"Writing...")
file_name = "WWHALigaData.xlsx"

with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
    df_transfers.to_excel(writer, sheet_name="Transfers", index=False)
    df_chips.to_excel(writer, sheet_name="Chip Usage", index=False)
    df_captaincy.to_excel(writer, sheet_name="Captaincy", index=False)

# === Step 5: Highlight TC usage in green ===
wb = load_workbook(file_name)
ws = wb["Captaincy"]
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):  # Triple Captain Used column
    for cell in row:
        if cell.value == "Yes":
            cell.fill = green_fill

wb.save(file_name)

print("✅ All done")
